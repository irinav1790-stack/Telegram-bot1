import telebot
import openpyxl
from flask import Flask, request
import os

EXCEL_FILE = 'препараты.xlsx'

def create_excel_if_not_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Препараты"
        ws.append(['Название препарата', 'Описание', 'Дозировка'])
        ws.append(['Парацетамол', 'Обезболивающее и жаропонижающее средство', '500 мг 3 раза в день'])
        wb.save(EXCEL_FILE)

create_excel_if_not_exists()

TOKEN = '8410859830:AAEqRSwrAXfUQaf4POcBRy9f_w9CS5_TNxE'
bot = telebot.TeleBot(TOKEN)
# ВАЖНО: Укажите здесь свой домен Render, без точки в конце!
WEBHOOK_URL = 'https://telegram-bot1.onrender.com/webhook'  # замените на ваш адрес Render

def read_drugs_from_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        drugs = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, description, dosage = row
            drugs.append({
                'name': name,
                'description': description,
                'dosage': dosage
            })
        return drugs
    except Exception:
        return []

@bot.message_handler(commands=['start'])
def send_welcome(message):
    bot.send_message(message.chat.id, "Здравствуйте! Я бот-справочник по препаратам. Введите название препарата или используйте команду /list для просмотра всех препаратов.")

@bot.message_handler(commands=['list'])
def list_drugs(message):
    drugs = read_drugs_from_excel()
    if not drugs:
        bot.send_message(message.chat.id, "Не удалось загрузить список препаратов.")
        return
    reply = "Список препаратов:\n"
    for drug in drugs:
        if drug['name']:
            reply += f"- {drug['name']}\n"
    bot.send_message(message.chat.id, reply)

@bot.message_handler(func=lambda message: True)
def search_drug(message):
    query = message.text.strip().lower()
    drugs = read_drugs_from_excel()
    found = None
    for drug in drugs:
        if drug['name'] and query in drug['name'].lower():
            found = drug
            break
    if found:
        reply = f"Название: {found['name']}\nОписание: {found['description']}\nДозировка: {found['dosage']}"
    else:
        reply = "Препарат не найден. Попробуйте другое название или используйте /list."
    bot.send_message(message.chat.id, reply)

app = Flask(__name__)

@app.route('/webhook', methods=['POST'])
def webhook():
    if request.method == 'POST':
        update = telebot.types.Update.de_json(request.stream.read().decode("utf-8"))
        bot.process_new_updates([update])
        return '', 200
    return 'Webhook endpoint', 200

@app.route('/', methods=['GET'])
def index():
    return 'Бот работает!'

def set_webhook():
    # Удаляем старый webhook и устанавливаем новый
    bot.remove_webhook()
    s = bot.set_webhook(url=WEBHOOK_URL)
    if s:
        print("Webhook установлен успешно.")
    else:
        print("Ошибка установки webhook.")

# --- Основная точка входа ---
if __name__ != '__main__':
    # Для Render и других WSGI серверов: всегда устанавливаем webhook при запуске
    set_webhook()
else:
    # Для локального запуска: используем polling, но выводим предупреждение
    print("ВНИМАНИЕ: Не используйте встроенный сервер Flask в production! Используйте gunicorn или другой WSGI сервер.")
    print("Локальный режим: используется polling, а не webhook.")
    bot.remove_webhook()
    bot.polling(none_stop=True)